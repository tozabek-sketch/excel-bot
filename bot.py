"""
Rasm → Excel Telegram Bot (yangilangan)
Rasm Excel ichiga haqiqiy rasm sifatida joylashtiriladi
"""

import os
import io
import logging
from pathlib import Path
from telegram import Update, Document
from telegram.ext import (
    ApplicationBuilder, CommandHandler, MessageHandler,
    filters, ContextTypes
)
from PIL import Image
import openpyxl
from openpyxl.drawing.image import Image as XLImage

try:
    import pillow_heif
    pillow_heif.register_heif_opener()
    HEIC_SUPPORTED = True
except ImportError:
    HEIC_SUPPORTED = False

logging.basicConfig(
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
    level=logging.INFO
)
logger = logging.getLogger(__name__)

TOKEN = "8951279462:AAEDjR6dMqi0BFBt5sSuUxmIdczCSP31u90"

SUPPORTED_EXTENSIONS = {".jpg", ".jpeg", ".png", ".bmp", ".webp", ".tiff", ".tif", ".heic", ".heif"}
MAX_FILE_SIZE = 20 * 1024 * 1024


def image_bytes_to_excel(img_bytes: bytes, filename: str = "image") -> bytes:
    """Rasmni Excel fayliga joylashtiradi (to'liq sifatda)"""
    img = Image.open(io.BytesIO(img_bytes))

    # RGBA → RGB
    if img.mode in ("RGBA", "P", "LA"):
        bg = Image.new("RGB", img.size, (255, 255, 255))
        bg.paste(img, mask=img.split()[-1] if img.mode in ("RGBA", "LA") else None)
        img = bg
    elif img.mode != "RGB":
        img = img.convert("RGB")

    orig_size = f"{img.width} × {img.height} px"

    # Excel uchun max 1500px
    max_dim = 1500
    if img.width > max_dim or img.height > max_dim:
        ratio = min(max_dim / img.width, max_dim / img.height)
        img = img.resize((int(img.width * ratio), int(img.height * ratio)), Image.LANCZOS)

    # PNG sifatida saqlash (siqilmagan)
    png_buf = io.BytesIO()
    img.save(png_buf, format="PNG", optimize=False)
    png_buf.seek(0)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Rasm"

    xl_img = XLImage(png_buf)
    xl_img.anchor = "B2"
    ws.add_image(xl_img)

    # Kataklarni rasmga moslashtirish
    col_count = max(img.width // 7 + 5, 20)
    row_count = max(img.height // 20 + 5, 20)
    for col in range(1, col_count):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 7
    for row in range(1, row_count):
        ws.row_dimensions[row].height = 20

    # Ma'lumot varag'i
    meta = wb.create_sheet("Ma'lumot")
    meta.column_dimensions["A"].width = 20
    meta.column_dimensions["B"].width = 30
    rows = [
        ("Fayl nomi", filename),
        ("Asl o'lcham", orig_size),
        ("Excel o'lcham", f"{img.width} × {img.height} px"),
        ("Format", "PNG (yuqori sifat)"),
        ("Bot", "@Imagesdanexcelga_bot"),
    ]
    for i, (k, v) in enumerate(rows, 1):
        meta.cell(row=i, column=1, value=k)
        meta.cell(row=i, column=2, value=v)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    heic = "✅ HEIC/HEIF" if HEIC_SUPPORTED else "⚠️ HEIC (kutubxona yo'q)"
    await update.message.reply_text(
        "👋 *Rasm → Excel Botga xush kelibsiz!*\n\n"
        "Menga rasm yuboring:\n"
        f"📷 JPG / JPEG / PNG\n📷 {heic}\n"
        "📷 BMP / WEBP / TIFF\n\n"
        "Men uni *yuqori sifatda Excel fayliga* (.xlsx) joylashtiraman! ✨",
        parse_mode="Markdown"
    )


async def process_photo(update: Update, context: ContextTypes.DEFAULT_TYPE):
    msg = await update.message.reply_text("⏳ Qayta ishlanmoqda...")
    photo = update.message.photo[-1]
    file = await context.bot.get_file(photo.file_id)
    img_bytes = await file.download_as_bytearray()
    try:
        excel_bytes = image_bytes_to_excel(bytes(img_bytes), "telegram_photo")
        await context.bot.send_document(
            chat_id=update.effective_chat.id,
            document=io.BytesIO(excel_bytes),
            filename="rasm.xlsx",
            caption="✅ Tayyor! Excel faylini oching — rasm ichida bo'ladi."
        )
    except Exception as e:
        logger.error(f"Photo error: {e}")
        await update.message.reply_text(f"❌ Xatolik: {e}")
    finally:
        await msg.delete()


async def process_document(update: Update, context: ContextTypes.DEFAULT_TYPE):
    doc: Document = update.message.document
    filename = doc.file_name or "file"
    ext = Path(filename).suffix.lower()

    if ext not in SUPPORTED_EXTENSIONS:
        await update.message.reply_text(f"⚠️ *{ext}* formati qo'llab-quvvatlanmaydi.", parse_mode="Markdown")
        return

    if not HEIC_SUPPORTED and ext in {".heic", ".heif"}:
        await update.message.reply_text("⚠️ HEIC hozir mavjud emas. JPG yoki PNG yuboring.")
        return

    if doc.file_size and doc.file_size > MAX_FILE_SIZE:
        await update.message.reply_text("❌ Fayl juda katta (maks. 20 MB).")
        return

    msg = await update.message.reply_text(f"⏳ *{filename}* qayta ishlanmoqda...", parse_mode="Markdown")
    try:
        file = await context.bot.get_file(doc.file_id)
        img_bytes = await file.download_as_bytearray()
        excel_bytes = image_bytes_to_excel(bytes(img_bytes), Path(filename).stem)
        out_name = Path(filename).stem + ".xlsx"
        await context.bot.send_document(
            chat_id=update.effective_chat.id,
            document=io.BytesIO(excel_bytes),
            filename=out_name,
            caption=f"✅ *{filename}* → *{out_name}* tayyor!",
            parse_mode="Markdown"
        )
    except Exception as e:
        logger.error(f"Document error: {e}")
        await update.message.reply_text(f"❌ Xatolik: {e}")
    finally:
        await msg.delete()


def main():
    app = ApplicationBuilder().token(TOKEN).build()
    app.add_handler(CommandHandler("start", start))
    app.add_handler(MessageHandler(filters.PHOTO, process_photo))
    app.add_handler(MessageHandler(filters.Document.ALL, process_document))
    logger.info("Bot ishga tushdi...")
    app.run_polling(drop_pending_updates=True)


if __name__ == "__main__":
    main()
